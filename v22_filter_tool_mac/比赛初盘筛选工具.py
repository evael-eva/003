# -*- coding: utf-8 -*-
"""
比赛初盘筛选工具 - 独立程序
从监控程序中提取逻辑，实现基于初盘的比賽筛选功能

功能：
1. 支持半场/全场亚盘和大小球初盘筛选
2. 每个筛选项包含：主水范围、盘口选择、客水范围
3. 使用浏览器进行盘口筛选，代理获取初盘数据
4. 导出符合条件的比赛为Excel文件

使用方法：
  python 比赛初盘筛选工具.py
"""

import sys
import os
import time
import random
import re
import threading
from datetime import datetime
from bs4 import BeautifulSoup
from DrissionPage import WebPage, ChromiumOptions
from PyQt5.QtCore import Qt, QTimer, QSize, pyqtSignal, QThread, QObject
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QPushButton, QTableWidget, QTableWidgetItem,
    QGroupBox, QTextEdit, QHeaderView, QMessageBox, QFrame, QSplitter,
    QListWidget, QListWidgetItem, QCheckBox, QSpinBox, QDoubleSpinBox,
    QLineEdit, QComboBox, QScrollArea, QProgressBar, QStatusBar,
    QDialog, QDateEdit, QFormLayout, QTabWidget,
    QAbstractItemView, QSizePolicy, QToolTip, QFileDialog
)
from PyQt5.QtGui import QFont, QColor, QBrush, QIcon, QPalette, QCursor
import requests
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor


# ============================================================
# 全局样式表（简洁风格）
# ============================================================
GLOBAL_STYLESHEET = """
    QMainWindow {
        background-color: #f5f5f5;
    }
    QWidget {
        font-family: "PingFang SC", "Microsoft YaHei", "Segoe UI", sans-serif;
        font-size: 13px;
        color: #333333;
    }
    QGroupBox {
        font-weight: bold;
        border: 1px solid #cccccc;
        border-radius: 6px;
        margin-top: 12px;
        padding-top: 10px;
        padding: 12px;
        background-color: white;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        left: 12px;
        padding: 0 8px;
        color: #2c3e50;
        font-weight: bold;
    }
    QPushButton {
        background-color: #3498db;
        color: white;
        border: none;
        padding: 10px 20px;
        border-radius: 5px;
        font-weight: bold;
        font-size: 13px;
        min-height: 28px;
    }
    QPushButton:hover {
        background-color: #2980b9;
    }
    QPushButton:pressed {
        background-color: #21618c;
    }
    QPushButton:disabled {
        background-color: #bdc3c7;
        color: #7f8c8d;
    }
    QPushButton#startBtn {
        background-color: #27ae60;
        font-size: 14px;
        padding: 12px 30px;
    }
    QPushButton#startBtn:hover { background-color: #229954; }
    QPushButton#exportBtn {
        background-color: #e67e22;
    }
    QPushButton#exportBtn:hover { background-color: #d35400; }
    
    QSpinBox, QDoubleSpinBox, QLineEdit, QComboBox {
        padding: 6px 10px;
        border: 1px solid #cccccc;
        border-radius: 4px;
        background-color: white;
        color: #333333;
        font-size: 13px;
    }
    QSpinBox:focus, QDoubleSpinBox:focus, QLineEdit:focus, QComboBox:focus {
        border: 1px solid #3498db;
    }
    
    QTableWidget {
        background-color: white;
        border: 1px solid #cccccc;
        border-radius: 6px;
        gridline-color: #e0e0e0;
        font-size: 12px;
    }
    QTableWidget::item {
        padding: 6px;
        color: #333333;
    }
    QTableWidget::item:selected {
        background-color: #ebf5fb;
        color: #2c3e50;
    }
    QHeaderView::section {
        background-color: #ecf0f1;
        color: #2c3e50;
        padding: 8px;
        border: none;
        border-bottom: 2px solid #bdc3c7;
        font-weight: bold;
        font-size: 12px;
    }
    
    QTextEdit {
        background-color: #fafafa;
        color: #555555;
        border: 1px solid #cccccc;
        border-radius: 4px;
        font-family: Consolas, "Courier New", monospace;
        font-size: 11px;
        padding: 8px;
    }
    
    QCheckBox {
        spacing: 6px;
        color: #333333;
        font-size: 12px;
    }
    QCheckBox::indicator {
        width: 18px;
        height: 18px;
        border-radius: 3px;
        border: 1px solid #cccccc;
        background-color: white;
    }
    QCheckBox::indicator:checked {
        background-color: #3498db;
        border-color: #3498db;
    }
"""


# ============================================================
# 嵌入的 FilterController 类（来自 v21/filter_controller.py）
# ============================================================
class FilterController:
    """盘口筛选控制器（非QObject，避免跨线程信号阻塞主UI）"""
    
    def __init__(self):
        self.web_page = None
        self._running = False
        self._log_callback = None  # 可选的外部日志回调
    
    def set_log_callback(self, callback):
        """设置日志回调函数，用于在工作线程中安全地输出日志"""
        self._log_callback = callback
    
    def _emit_log(self, msg):
        """通过回调输出日志，如果无回调则print"""
        if self._log_callback:
            try:
                self._log_callback(msg)
            except:
                print(msg)
        else:
            print(msg)

    ASIAN_HANDICAP_OPTIONS = {
        '-0.25': '平/半', '-0.5': '半球', '-0.75': '半/一', '-1': '一球',
        '-1.25': '一/球半', '-1.5': '球半', '-1.75': '球半/两', '-2': '两球',
        '-2.25': '两/两球半', '-2.75': '两球半/三', '-3': '三球',
        '-3.5': '三球半', '-5.25': '五/五球半',
        '0': '平手', '0.25': '平/半', '0.5': '半球', '0.75': '半/一',
        '1': '一球', '1.25': '一/球半', '1.5': '球半', '1.75': '球半/两',
        '2': '两球', '2.5': '两球半', '2.75': '两球半/三', '3.25': '三/三球半',
        '3.5': '三球半', '4.5': '四球半',
    }

    OVERUNDER_OPTIONS = {
        '1.75': '1.5/2', '2': '2', '2.25': '2/2.5', '2.5': '2.5',
        '2.75': '2.5/3', '3': '3', '3.25': '3/3.5', '3.5': '3.5',
        '3.75': '3.5/4', '4': '4', '4.25': '4/4.5', '4.5': '4.5',
        '4.75': '4.5/5', '5.25': '5/5.5', '6': '6',
    }

    # __init__ 已在类定义中，此处删除重复

    @staticmethod
    def get_chrome_path():
        """获取Chrome浏览器路径 - 跨平台兼容"""
        import sys as _sys
        
        if _sys.platform == 'darwin':
            # macOS Chrome 路径
            mac_paths = [
                '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome',
                '/Applications/Chromium.app/Contents/MacOS/Chromium',
                '/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge',
                os.path.expanduser('~/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'),
            ]
            for path in mac_paths:
                if os.path.exists(path):
                    return path
            # 尝试用 mdfind 查找
            try:
                import subprocess
                result = subprocess.run(
                    ['mdfind', 'kMDItemCFBundleIdentifier == "com.google.Chrome"'],
                    capture_output=True, text=True, timeout=5
                )
                if result.stdout.strip():
                    chrome_app = result.stdout.strip().split('\n')[0]
                    chrome_path = os.path.join(chrome_app, 'Contents', 'MacOS', 'Google Chrome')
                    if os.path.exists(chrome_path):
                        return chrome_path
            except:
                pass
            return None
        elif _sys.platform == 'win32':
            possible_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                r"D:\Program Files\Google\Chrome\Application\chrome.exe",
                f"{os.getcwd()}\\images\\Chrome\\chrome.exe"
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    return path
            try:
                import winreg
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                     r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe")
                path = winreg.QueryValue(key, None)
                winreg.CloseKey(key)
                if os.path.exists(path):
                    return path
            except:
                pass
            return None
        else:
            # Linux
            linux_paths = [
                '/usr/bin/google-chrome',
                '/usr/bin/chromium-browser',
                '/usr/bin/chromium',
                '/snap/bin/chromium',
            ]
            for path in linux_paths:
                if os.path.exists(path):
                    return path
            return None

    @staticmethod
    def _random_delay(min_s=0.5, max_s=1.5):
        time.sleep(random.uniform(min_s, max_s))

    def create_browser(self, headless=False):
        co = ChromiumOptions()
        chrome_path = self.get_chrome_path()
        co.headless(headless)
        co.auto_port(True)
        co.no_js(False)
        co.mute(True)
        user_agents = [
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
        ]
        co.set_user_agent(random.choice(user_agents))
        if chrome_path:
            co.set_browser_path(chrome_path)
        self.web_page = WebPage('d', chromium_options=co)

    def close_browser(self):
        if self.web_page:
            try:
                self.web_page.quit()
            except:
                pass
            finally:
                self.web_page = None

    def open_live_page(self):
        if not self.web_page:
            self.create_browser(headless=True)
        url = "https://live.titan007.com/indexall.aspx"
        self.web_page.get(url, timeout=30)
        self._random_delay(2, 3)
        try:
            erheyi_btn = self.web_page.ele('xpath://*[@id="tools"]/ul/li[1]', timeout=5)
            if erheyi_btn:
                erheyi_btn.click()
                self._random_delay(0.5, 1)
        except Exception as e:
            print(f"[FilterController] 点击二合一按钮失败: {e}")
        try:
            complete_btn = self.web_page.ele('xpath://*[@id="button6"]', timeout=5)
            if complete_btn:
                complete_btn.click()
                self._random_delay(0.5, 1)
        except Exception as e:
            print(f"[FilterController] 点击完整按钮失败: {e}")

    def _click_goal_div_button(self):
        try:
            btn = self.web_page.ele('xpath://*[@id="button8"]', timeout=5)
            if btn:
                btn.click()
                time.sleep(2)
                print(f"[FilterController] 打开弹窗: ok")
                return True
        except Exception as e:
            print(f"[FilterController] 点击盘*按钮失败: {e}")
        return False

    def _ensure_asian_tab(self):
        try:
            radio = self.web_page.ele('xpath://input[@id="radioGoalType0"]', timeout=5)
            if radio:
                radio.click()
                time.sleep(1)
                print(f"[FilterController] 切换到亚让tab: ok")
                return True
        except Exception as e:
            print(f"[FilterController] 切换到亚让tab失败: {e}")
        return False

    def _switch_to_overunder_tab(self):
        try:
            radio = self.web_page.ele('xpath://input[@id="radioGoalType1"]', timeout=5)
            if radio:
                radio.click()
                time.sleep(1)
                print(f"[FilterController] 切换到进球数tab: ok")
                return True
        except Exception as e:
            print(f"[FilterController] 切换到进球数tab失败: {e}")
        return False

    def _uncheck_all_checkboxes(self):
        try:
            checkboxes = self.web_page.eles(
                'xpath://div[@id="goalDiv"]//input[@type="checkbox" and @name="checkbox"]',
                timeout=5
            )
            unchecked = 0
            for cb in checkboxes:
                try:
                    if cb.states.is_checked:
                        cb.click()
                        unchecked += 1
                        time.sleep(0.05)
                except:
                    continue
            print(f"[FilterController] 兜底检查: {len(checkboxes)}个checkbox, 其中{unchecked}个已选中需取消")
        except Exception as e:
            print(f"[FilterController] 取消全选失败: {e}")

    def _check_selected_handicaps(self, selected_values):
        count = 0
        if not selected_values:
            print(f"[FilterController] 勾选盘口: selected_values为空")
            return count
        selected_set = set(selected_values)
        try:
            checkboxes = self.web_page.eles(
                'xpath://div[@id="goalDiv"]//input[@type="checkbox" and @name="checkbox"]',
                timeout=5
            )
            print(f"[FilterController] 勾选盘口: 找到{len(checkboxes)}个checkbox, 目标值={list(selected_set)}")
            page_samples = []
            for cb in checkboxes[:10]:
                try:
                    page_samples.append(cb.attr('value'))
                except:
                    page_samples.append('?')
            print(f"[FilterController] 页面值样本={page_samples}")
            matched = set()
            for cb in checkboxes:
                try:
                    val = cb.attr('value')
                    if val and val in selected_set:
                        matched.add(val)
                        if not cb.states.is_checked:
                            cb.click()
                            count += 1
                            time.sleep(0.15)
                except Exception:
                    continue
            not_found = selected_set - matched
            if not_found:
                print(f"[FilterController] 未找到的值={not_found}")
            print(f"[FilterController] 实际勾选{count}个")
            time.sleep(0.3)
        except Exception as e:
            print(f"[FilterController] 勾选盘口失败: {e}")
        return count

    def _click_confirm(self):
        try:
            confirm_btn = self.web_page.ele(
                'xpath://div[@id="selectGoals_div"]//input[@type="button" and @value="确定"]',
                timeout=5
            )
            if confirm_btn:
                confirm_btn.click()
                time.sleep(1)
                print(f"[FilterController] 点击确定: ok")
                return True
        except Exception as e:
            print(f"[FilterController] 点击确定失败: {e}")
        return False

    def filter_by_asian(self, selected_values, exclude_finished=False):
        matches = []
        try:
            self._click_goal_div_button()
            self._ensure_asian_tab()
            self._uncheck_all_checkboxes()
            checked_count = self._check_selected_handicaps(selected_values)
            print(f"[FilterController] 亚盘筛选: 已勾选 {checked_count} 个盘口")
            self._click_confirm()
            matches = self._parse_match_list(source='亚盘', exclude_finished=exclude_finished)
        except Exception as e:
            print(f"[FilterController] 亚盘筛选异常: {e}")
        return matches

    def filter_by_overunder(self, selected_values, exclude_finished=False):
        matches = []
        try:
            self._click_goal_div_button()
            self._switch_to_overunder_tab()
            self._uncheck_all_checkboxes()
            checked_count = self._check_selected_handicaps(selected_values)
            print(f"[FilterController] 大小球筛选: 已勾选 {checked_count} 个盘口")
            self._click_confirm()
            matches = self._parse_match_list(source='大小球', exclude_finished=exclude_finished)
        except Exception as e:
            print(f"[FilterController] 大小球筛选异常: {e}")
        return matches
    
    def get_all_matches(self, exclude_finished=False):
        matches = []
        try:
            matches = self._parse_match_list(source='全部', exclude_finished=exclude_finished)
            self._emit_log(f"[FilterController] 获取到 {len(matches)} 场比赛")
        except Exception as e:
            self._emit_log(f"[FilterController] 获取所有比赛失败: {e}")
        return matches

    def filter_combined(self, asian_enabled, asian_values, ou_enabled, ou_values,
                       half_ou_enabled=False, half_ou_min=0.75, half_ou_max=1.5,
                       skip_finished=False):
        all_matches = {}
        if half_ou_enabled and not asian_enabled and not ou_enabled:
            self._emit_log("[FilterController] 检测到仅启用半场大小球筛选，获取所有比赛...")
            self.open_live_page()
            all_matches_list = self._parse_match_list(source='全部', exclude_finished=True)
            self._emit_log(f"[FilterController] 获取到 {len(all_matches_list)} 场未结束比赛")
            for m in all_matches_list:
                mid = m.get('match_id', '')
                if mid and mid not in all_matches:
                    m['source'] = '全部'
                    all_matches[mid] = m
            return list(all_matches.values())
        browser_filters = []
        if asian_enabled and asian_values:
            browser_filters.append('asian')
        if ou_enabled and ou_values:
            browser_filters.append('ou')
        need_parallel = len(browser_filters) >= 2
        if need_parallel:
            print("[FilterController] 启动双浏览器并行筛选...")
            asian_results = []
            ou_results = []
            asian_error = None
            ou_error = None
            def _run_asian():
                nonlocal asian_results, asian_error
                try:
                    ctrl_a = FilterController()
                    ctrl_a.open_live_page()
                    asian_results = ctrl_a.filter_by_asian(asian_values, exclude_finished=skip_finished)
                    ctrl_a.close_browser()
                except Exception as e:
                    asian_error = str(e)
            def _run_ou():
                nonlocal ou_results, ou_error
                try:
                    ctrl_o = FilterController()
                    ctrl_o.open_live_page()
                    ou_results = ctrl_o.filter_by_overunder(ou_values, exclude_finished=skip_finished)
                    ctrl_o.close_browser()
                except Exception as e:
                    ou_error = str(e)
            t1 = threading.Thread(target=_run_asian)
            t2 = threading.Thread(target=_run_ou)
            t1.start()
            t2.start()
            t1.join(timeout=45)  # v22修复: 减少超时避免长时间阻塞
            t2.join(timeout=45)
            if asian_error:
                print(f"[FilterController] 亚盘线程异常: {asian_error}")
            if ou_error:
                print(f"[FilterController] 大小球线程异常: {ou_error}")
            for m in asian_results:
                mid = m.get('match_id', '')
                if mid and mid not in all_matches:
                    m['source'] = '亚盘'
                    all_matches[mid] = m
                elif mid and mid in all_matches:
                    existing = all_matches[mid]
                    sources = existing.get('sources', [existing.get('source', '')])
                    if '亚盘' not in sources:
                        sources.append('亚盘')
                    existing['sources'] = sources
            for m in ou_results:
                mid = m.get('match_id', '')
                if mid and mid not in all_matches:
                    m['source'] = '大小球'
                    all_matches[mid] = m
                elif mid and mid in all_matches:
                    existing = all_matches[mid]
                    sources = existing.get('sources', [existing.get('source', '')])
                    if '大小球' not in sources:
                        sources.append('大小球')
                    existing['sources'] = sources
        else:
            self.open_live_page()
            if asian_enabled and asian_values:
                asian_matches = self.filter_by_asian(asian_values, exclude_finished=skip_finished)
                for m in asian_matches:
                    mid = m.get('match_id', '')
                    if mid and mid not in all_matches:
                        m['source'] = '亚盘'
                        all_matches[mid] = m
                    elif mid and mid in all_matches:
                        existing = all_matches[mid]
                        sources = existing.get('sources', [existing.get('source', '')])
                        if '亚盘' not in sources:
                            sources.append('亚盘')
                        existing['sources'] = sources
            if ou_enabled and ou_values:
                ou_matches = self.filter_by_overunder(ou_values, exclude_finished=skip_finished)
                for m in ou_matches:
                    mid = m.get('match_id', '')
                    if mid and mid not in all_matches:
                        m['source'] = '大小球'
                        all_matches[mid] = m
                    elif mid and mid in all_matches:
                        existing = all_matches[mid]
                        sources = existing.get('sources', [existing.get('source', '')])
                        if '大小球' not in sources:
                            sources.append('大小球')
                        existing['sources'] = sources
            if not asian_enabled and not ou_enabled and half_ou_enabled:
                self._emit_log("[FilterController] 仅启用半场大球初盘筛选，获取所有比赛...")
                all_matches_list = self.get_all_matches()
                for m in all_matches_list:
                    mid = m.get('match_id', '')
                    if mid and mid not in all_matches:
                        m['source'] = '全部'
                        all_matches[mid] = m
        return list(all_matches.values())

    def _extract_minutes(self, status):
        try:
            if not status:
                return 0
            if status.isdigit():
                return int(status)
            elif '中' in status:
                return 45
            else:
                match_minute = re.search(r'(\d+)', status)
                if match_minute:
                    return int(match_minute.group(1))
            return 0
        except:
            return 0

    def _is_match_live(self, status):
        if not status or not status.strip():
            return False
        s = status.strip()
        if s.isdigit():
            return True
        if s == '中':
            return True
        return False

    def _parse_match_list(self, source='未指定', exclude_finished=False):
        matches = []
        try:
            page_html = self.web_page.html
            soup = BeautifulSoup(page_html, 'html.parser')
            match_rows = soup.find_all('tr', id=lambda x: x and x.startswith('tr1_'))
            for row in match_rows:
                style = row.get('style', '')
                if 'display: none' in style:
                    continue
                tds = row.find_all('td')
                if len(tds) < 11:
                    continue
                match_id = ''
                row_id = row.get('id', '')
                if row_id and row_id.startswith('tr1_'):
                    match_id = row_id.replace('tr1_', '')
                league = tds[1].get_text(strip=True) if len(tds) > 1 else ''
                match_time = tds[2].get_text(strip=True) if len(tds) > 2 else ''
                status = tds[3].get_text(strip=True) if len(tds) > 3 else ''
                home_team = tds[4].get_text(strip=True) if len(tds) > 4 else ''
                score = ''
                if len(tds) > 5:
                    score_fonts = tds[5].find_all('font')
                    if len(score_fonts) >= 2:
                        score = f"{score_fonts[0].get_text(strip=True)}-{score_fonts[1].get_text(strip=True)}"
                    elif '-' in tds[5].get_text(strip=True):
                        score = tds[5].get_text(strip=True)
                away_team = tds[6].get_text(strip=True) if len(tds) > 6 else ''
                half_score = ''
                corner_data = ''
                if len(tds) > 7:
                    corner_span = tds[7].find('span', class_='td_halfB')
                    if corner_span:
                        corner_data = corner_span.get_text(strip=True)
                    half_span = tds[7].find('span', class_='td_halfR')
                    if half_span:
                        half_score = half_span.get_text(strip=True)
                handicap = ''
                if len(tds) > 10:
                    div_elem = tds[10].find('div')
                    if div_elem:
                        handicap = div_elem.get_text(strip=True)
                    else:
                        handicap = tds[10].get_text(strip=True)
                overunder = ''
                if len(tds) > 11:
                    ou_div = tds[11].find('div')
                    if ou_div:
                        overunder = ou_div.get_text(strip=True)
                    else:
                        overunder = tds[11].get_text(strip=True)
                if exclude_finished:
                    ended_statuses = ['完', '完场', '结束', '已结束', 'FIN', 'FT']
                    if status in ended_statuses:
                        continue
                if match_id:
                    matches.append({
                        'match_id': match_id,
                        'league': league,
                        'match_time': match_time,
                        'status': status,
                        'home_team': home_team,
                        'score': score,
                        'away_team': away_team,
                        'half_score': half_score,
                        'corner_data': corner_data,
                        'handicap': handicap,
                        'overunder': overunder,
                        'source': source,
                    })
        except Exception as e:
            print(f"[FilterController] 解析比赛列表失败: {e}")
        return matches


# ============================================================
# 嵌入的 OddsFetcher 类（来自 v21/odds_fetcher.py）
# ============================================================
USER_AGENT_POOL = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36 Edg/147.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36 Edg/146.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:132.0) Gecko/20100101 Firefox/132.0',
]

DEFAULT_HEADERS_TEMPLATE = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
    'Cache-Control': 'no-cache',
    'Pragma': 'no-cache',
    'Priority': 'u=0, i',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
}

DEFAULT_COOKIE = (
    'b-user-id=a7f93b80-e642-0d71-94d1-1cc4576514c6; '
    'letgoalSettingSolution=; detailCookie=null; ishidepcad=; '
    'Hm_lvt_a88664a99dbcb9c7c07dc420114041b3=1773720401,1775723609; '
    'HMACCOUNT=DC1D86079E63CA12; '
    'Hm_lvt_3c285e4976a3c4fb2124b4d51dd1801e=1774270674,1775730459; '
    'totalSettingSolution=; '
    'Hm_lpvt_3c285e4976a3c4fb2124b4d51dd1801e=1776258084; '
    'Hm_lpvt_a88664a99dbcb9c7c07dc420114041b3=1776258394; '
    'xallpubLinkMatch=4f450e0644019ed94385423f0ce7ea1c157b9f21df0fb08df4bb91d57a5cc808'
    '4b024934ca5b6156b431345d9f235f950e6ef711de1ad4ea1b2e8e66f16e73ccfffbd80f53b1f2ed73850715dd19a5d704dc7fd9d1992b62a0c5f605bd89a271bb31123c8dbc8c2e8ec000e05ab07806101390c2ebc9d9cbe378bf3b73b8ae80b0350d14adc3aed05be42dc4b432be945686cab789b2191862bbd0a4a05e9865153af293e1e8b3f0536b27de15217ab66fa5d8fb83d51428b7b3cc8f9e36ac4b'
)


class OddsFetcher:
    """盘口数据获取器"""

    def __init__(self, proxy_config=None):
        self.session = requests.Session()
        from requests.adapters import HTTPAdapter
        adapter = HTTPAdapter(pool_connections=30, pool_maxsize=30, max_retries=3)
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)
        self.session.headers.update(DEFAULT_HEADERS_TEMPLATE)
        if DEFAULT_COOKIE:
            self.session.headers.update({'Cookie': DEFAULT_COOKIE})
        self.company_id = 3
        self.proxy_config = proxy_config
        self.current_proxy = None
        self.proxies = None
        self.proxy_start_time = None
        self.PROXY_ROTATE_INTERVAL = 600
        self._request_count = 0
        self._proxy_lock = threading.RLock()  # v22修复: 可重入锁，避免_mark_proxy_failed调用_fetch_new_proxy时死锁
        self._last_fetch_time = 0
        self._fetch_cooldown = 2
        
        # v24新增: 代理失效管理
        self.failed_proxies = set()  # 失效代理黑名单
        self.failed_proxy_count = 0  # 连续失败次数
        self.max_failed_before_rotate = 3  # 连续失败3次后强制轮换
        if proxy_config and proxy_config.get('enabled'):
            self._setup_single_proxy(proxy_config)
    
    @staticmethod
    def _get_random_user_agent():
        return random.choice(USER_AGENT_POOL)
    
    def _setup_single_proxy(self, proxy_config):
        try:
            api_url = proxy_config.get('api_url', '')
            if not api_url:
                return
            proxies_list = self._extract_proxies_from_api(api_url, num=1)
            if not proxies_list:
                return
            self.current_proxy = proxies_list[0]
            self._apply_current_proxy()
            self.proxy_start_time = time.time()
        except Exception as e:
            print(f"[代理初始化] ❌ 设置失败: {e}")
            self.current_proxy = None
            self.proxies = None
    
    def _extract_proxies_from_api(self, api_url, num=10):
        try:
            api_url = api_url.replace('\\r\\n', '').replace('\\n', '').replace('\\r', '')
            if 'num=' not in api_url:
                if '?' in api_url:
                    api_url += f"&num={num}"
                else:
                    api_url += f"?num={num}"
            else:
                import re as re_module
                api_url = re_module.sub(r'num=\d+', f'num={num}', api_url)
            response = requests.get(api_url, timeout=8)  # v22修复: 8秒超时避免长时间卡住
            if response.status_code != 200:
                return []
            text = response.text.strip()
            if not text:
                return []
            import re as re_module
            pattern = r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}):(\d{2,5})'
            matches = re_module.findall(pattern, text)
            proxies = [f'{ip}:{port}' for ip, port in matches]
            valid_proxies = list(set(proxies))
            return valid_proxies
        except Exception as e:
            return []
    
    def _apply_current_proxy(self):
        if not self.current_proxy:
            self.proxies = None
            return
        auth_username = self.proxy_config.get('auth_username', '') if self.proxy_config else ''
        auth_password = self.proxy_config.get('auth_password', '') if self.proxy_config else ''
        if auth_username and auth_password:
            proxy_url = f"http://{auth_username}:{auth_password}@{self.current_proxy}"
        else:
            proxy_url = f"http://{self.current_proxy}"
        self.proxies = {"http": proxy_url, "https": proxy_url}
    
    def _rotate_proxy_if_needed(self):
        """v24修复: 定时轮换代理，同时清理过期黑名单"""
        if not self.current_proxy or not self.proxy_start_time:
            return False
        elapsed = time.time() - self.proxy_start_time
        if elapsed >= self.PROXY_ROTATE_INTERVAL:
            # v24新增: 轮换前清理过期的黑名单（超过30分钟的失效记录）
            self._cleanup_failed_proxies()
            self._fetch_new_proxy()
            return True
        return False
    
    def _cleanup_failed_proxies(self):
        """v24新增: 清理过期的失效代理记录"""
        # 如果黑名单太大（超过50个），清空一半
        if len(self.failed_proxies) > 50:
            # 保留最近的25个，清空其他的
            proxies_list = list(self.failed_proxies)
            self.failed_proxies = set(proxies_list[-25:])
            print(f"[代理管理] 🧹 清理黑名单: 从{len(proxies_list)}个减少到{len(self.failed_proxies)}个")
    
    def _fetch_new_proxy(self, force=False):
        """v24修复: 获取新代理，支持强制轮换和黑名单过滤"""
        if not self.proxy_config or not self.proxy_config.get('enabled'):
            return False
        with self._proxy_lock:
            current_time = time.time()
            elapsed = current_time - self._last_fetch_time
            
            # v24新增: 如果不是强制轮换，检查冷却时间
            if not force and elapsed < self._fetch_cooldown:
                return False
            
            api_url = self.proxy_config.get('api_url', '')
            
            # v24新增: 尝试多次获取，直到获得不在黑名单中的代理
            max_attempts = 3
            for attempt in range(max_attempts):
                proxies_list = self._extract_proxies_from_api(api_url, num=5)  # 一次获取多个
                
                if not proxies_list:
                    print(f"[代理管理] ❌ API返回为空 (尝试{attempt+1}/{max_attempts})")
                    continue
                
                # v24新增: 过滤黑名单中的代理
                valid_proxies = [p for p in proxies_list if p not in self.failed_proxies]
                
                if valid_proxies:
                    # 选择第一个有效代理
                    new_proxy = valid_proxies[0]
                    self.current_proxy = new_proxy
                    self._apply_current_proxy()
                    self.proxy_start_time = time.time()
                    self._last_fetch_time = time.time()
                    self.failed_proxy_count = 0  # 重置失败计数
                    
                    print(f"[代理管理] ✅ 切换到新代理: {new_proxy} (黑名单中有{len(self.failed_proxies)}个失效代理)")
                    return True
                else:
                    print(f"[代理管理] ⚠️ 所有代理都在黑名单中，等待冷却后重试 (尝试{attempt+1}/{max_attempts})")
                    if attempt < max_attempts - 1:
                        time.sleep(1)  # 等待1秒后重试
            
            # 如果所有尝试都失败，记录时间但不切换
            self._last_fetch_time = time.time()
            print(f"[代理管理] ❌ 无法获取有效代理，继续使用当前代理")
            return False
    
    def _mark_proxy_failed(self, proxy_addr):
        """v24修复: 标记代理失效，并加入黑名单
        
        v24优化: 只有连续失败达到阈值时才轮换代理，避免频繁更换
        """
        # 提取纯IP:Port格式
        if proxy_addr.startswith('http://'):
            proxy_addr = proxy_addr[7:]
        if '@' in proxy_addr:
            proxy_addr = proxy_addr.split('@')[1]
        
        need_rotate = False
        with self._proxy_lock:
            # v24新增: 加入失效黑名单
            self.failed_proxies.add(proxy_addr)
            self.failed_proxy_count += 1
            
            print(f"[代理管理] ❌ 代理 {proxy_addr} 失效 (累计失败{self.failed_proxy_count}次)")
            
            # v24优化: 只有连续失败3次后才强制轮换，避免频繁更换
            if self.failed_proxy_count >= self.max_failed_before_rotate:
                print(f"[代理管理] ⚠️ 连续失败{self.failed_proxy_count}次，强制轮换代理")
                self.failed_proxy_count = 0  # 重置计数器
                need_rotate = True
            # v24优化: 否则不立即更换，继续使用当前代理（其他线程可能成功）
            else:
                print(f"[代理管理] ℹ️ 失败次数未达阈值({self.max_failed_before_rotate})，暂不轮换")
        
        # v22修复: 释放锁后再执行网络请求，避免阻塞其他线程
        if need_rotate:
            self._fetch_new_proxy(force=True)  # 强制获取新代理

    def fetch_initial_odds(self, match_id):
        result = {'asian_initial': None, 'ou_initial': None, 'asian_rows': [], 'ou_rows': [], 'error': None}
        asian_data = self.fetch_asian_odds(match_id)
        if asian_data.get('error'):
            result['error'] = f"亚盘: {asian_data['error']}"
        else:
            result['asian_rows'] = asian_data.get('rows', [])
            if asian_data['rows']:
                initial_row = asian_data['rows'][-1]
                result['asian_initial'] = self._extract_asian_row(initial_row)
        ou_data = self.fetch_overunder_odds(match_id)
        if ou_data.get('error'):
            err = result.get('error', '')
            result['error'] = f"{err}; 大小球: {ou_data['error']}" if err else f"大小球: {ou_data['error']}"
        else:
            result['ou_rows'] = ou_data.get('rows', [])
            if ou_data['rows']:
                initial_row = ou_data['rows'][-1]
                result['ou_initial'] = self._extract_ou_row(initial_row)
        time.sleep(random.uniform(0.1, 0.2))
        return result

    def fetch_asian_odds(self, match_id):
        url = f"https://vip.titan007.com/changeDetail/handicap.aspx?id={match_id}&companyID={self.company_id}&l=0"
        headers = dict(self.session.headers)
        headers['User-Agent'] = self._get_random_user_agent()
        headers['Referer'] = f'https://vip.titan007.com/AsianOdds_n.aspx?id={match_id}&l=0'
        self._rotate_proxy_if_needed()
        current_proxy = self.current_proxy if self.current_proxy else '无代理'
        try:
            response = self.session.get(url, timeout=15, headers=headers, proxies=self.proxies)
            if response.status_code != 200:
                return {'header': [], 'rows': [], 'error': f'HTTP {response.status_code}'}
            html_content = self._decode_response(response)
            return self._parse_asian_table(html_content, f"比赛{match_id}")
        except requests.exceptions.ProxyError as e:
            self._mark_proxy_failed(current_proxy)
            return {'header': [], 'rows': [], 'error': f'代理错误: {str(e)}'}
        except requests.exceptions.Timeout:
            return {'header': [], 'rows': [], 'error': '请求超时'}
        except Exception as e:
            return {'header': [], 'rows': [], 'error': str(e)}

    def fetch_overunder_odds(self, match_id):
        url = f"https://vip.titan007.com/changeDetail/overunder.aspx?id={match_id}&companyID={self.company_id}&l=0"
        headers = dict(self.session.headers)
        headers['User-Agent'] = self._get_random_user_agent()
        headers['Referer'] = f'https://vip.titan007.com/AsianOdds_n.aspx?id={match_id}&l=0'
        self._rotate_proxy_if_needed()
        current_proxy = self.current_proxy if self.current_proxy else '无代理'
        try:
            response = self.session.get(url, timeout=15, headers=headers, proxies=self.proxies)
            if response.status_code != 200:
                return {'header': [], 'rows': [], 'error': f'HTTP {response.status_code}'}
            html_content = self._decode_response(response)
            return self._parse_overunder_table(html_content, f"比赛{match_id}")
        except requests.exceptions.ProxyError as e:
            self._mark_proxy_failed(current_proxy)
            return {'header': [], 'rows': [], 'error': f'代理错误: {str(e)}'}
        except requests.exceptions.Timeout:
            return {'header': [], 'rows': [], 'error': '请求超时'}
        except Exception as e:
            return {'header': [], 'rows': [], 'error': str(e)}

    def fetch_half_time_initial(self, match_id):
        """
        v24修复: 获取半场亚盘和大小球初盘数据（使用 changeDetail/handicapHalf.aspx 和 overunderHalf.aspx）
        :param match_id: 比赛ID
        :return: {
            'asian_half_initial': {'home_odds', 'handicap', 'away_odds', ...},
            'ou_half_initial': {'over_odds', 'goal_line', 'under_odds', ...},
            'error': str or None
        }
        """
        result = {
            'asian_half_initial': None,
            'ou_half_initial': None,
            'error': None,
        }

        # 获取半场亚盘
        asian_half_url = f"https://vip.titan007.com/changeDetail/handicapHalf.aspx?id={match_id}&companyID={self.company_id}&h=1&l=0"
        headers = dict(self.session.headers)
        headers['User-Agent'] = self._get_random_user_agent()
        headers['Referer'] = f'https://vip.titan007.com/AsianOdds_n.aspx?id={match_id}&l=0'
        
        self._rotate_proxy_if_needed()
        current_proxy = self.current_proxy if self.current_proxy else '无代理'
        
        try:
            response = self.session.get(asian_half_url, timeout=15, headers=headers, proxies=self.proxies)
            if response.status_code == 200:
                html_content = self._decode_response(response)
                parsed = self._parse_asian_table(html_content, f"半场亚盘{match_id}")
                if not parsed.get('error') and parsed.get('rows'):
                    # 取最后一行作为初盘
                    initial_row = parsed['rows'][-1]
                    result['asian_half_initial'] = self._extract_asian_row(initial_row)
            else:
                result['error'] = f"半场亚盘 HTTP {response.status_code}"
        except requests.exceptions.ProxyError as e:
            self._mark_proxy_failed(current_proxy)
            result['error'] = f"半场亚盘代理错误: {str(e)}"
        except Exception as e:
            result['error'] = f"半场亚盘: {str(e)}"

        # 短暂延迟
        time.sleep(random.uniform(0.1, 0.15))

        # 获取半场大小球
        ou_half_url = f"https://vip.titan007.com/changeDetail/overunderHalf.aspx?id={match_id}&companyID={self.company_id}&h=1&l=0"
        
        try:
            response = self.session.get(ou_half_url, timeout=15, headers=headers, proxies=self.proxies)
            if response.status_code == 200:
                html_content = self._decode_response(response)
                parsed = self._parse_overunder_table(html_content, f"半场大小球{match_id}")
                if not parsed.get('error') and parsed.get('rows'):
                    # 取最后一行作为初盘
                    initial_row = parsed['rows'][-1]
                    result['ou_half_initial'] = self._extract_ou_row(initial_row)
            else:
                err = result.get('error', '')
                result['error'] = f"{err}; 半场大小球 HTTP {response.status_code}" if err else f"半场大小球 HTTP {response.status_code}"
        except requests.exceptions.ProxyError as e:
            self._mark_proxy_failed(current_proxy)
            err = result.get('error', '')
            result['error'] = f"{err}; 半场大小球代理错误: {str(e)}" if err else f"半场大小球代理错误: {str(e)}"
        except Exception as e:
            err = result.get('error', '')
            result['error'] = f"{err}; 半场大小球: {str(e)}" if err else f"半场大小球: {str(e)}"

        # 短暂延迟
        time.sleep(random.uniform(0.1, 0.15))

        return result

    @staticmethod
    def _decode_response(response):
        raw_bytes = response.content
        content_type = response.headers.get('content-type', '').lower()
        if 'gb2312' in content_type or 'gbk' in content_type:
            return raw_bytes.decode('gb2312', errors='ignore')
        if 'utf-8' in content_type or 'utf8' in content_type:
            return raw_bytes.decode('utf-8', errors='ignore')
        encodings_to_try = ['utf-8', 'gb2312', 'gbk', 'iso-8859-1', 'big5']
        for enc in encodings_to_try:
            try:
                text = raw_bytes.decode(enc, errors='strict')
                if len(text) > 50 and ('<' in text or '\u4e00' <= ''.join(filter(str.isalpha, text[:100]))):
                    return text
                decoded = raw_bytes.decode(enc, errors='ignore')
                if decoded.strip():
                    return decoded
            except (UnicodeDecodeError, LookupError):
                continue
        return raw_bytes.decode('gb2312', errors='ignore')

    @staticmethod
    def _extract_asian_row(row_data):
        if len(row_data) >= 5:
            return {
                'time': row_data[0] if len(row_data) > 0 else '',
                'score': row_data[1] if len(row_data) > 1 else '',
                'home_odds': row_data[2] if len(row_data) > 2 else '',
                'handicap': row_data[3] if len(row_data) > 3 else '',
                'away_odds': row_data[4] if len(row_data) > 4 else '',
                'change_time': row_data[5] if len(row_data) > 5 else '',
                'status': row_data[6] if len(row_data) > 6 else '',
            }
        return {}

    @staticmethod
    def _convert_handicap_to_number(handicap_str):
        """
        v24修复: 将盘口字符串转换为数字（支持中文盘口）
        :param handicap_str: 盘口字符串（如"平手"、"平/半"、"1.25"等）
        :return: 转换后的浮点数，失败返回None
        """
        if not handicap_str or not isinstance(handicap_str, str):
            return None
        
        handicap_str = handicap_str.strip()
        
        # v24新增: 中文盘口映射表
        chinese_handicap_map = {
            # 正数盘口
            '平手': 0,
            '平/半': 0.25,
            '平半': 0.25,
            '半球': 0.5,
            '半/一': 0.75,
            '半一': 0.75,
            '一球': 1,
            '一/球半': 1.25,
            '一球/球半': 1.25,  # v24补充: 别名
            '一球半': 1.5,
            '球半': 1.5,
            '一/半': 1.25,
            '两球': 2,
            '二球': 2,
            '两/半': 2.25,
            '二/半': 2.25,
            '两球半': 2.5,
            '二球半': 2.5,
            '三球': 3,
            '三/半': 3.25,
            '三球半': 3.5,
            
            # 负数盘口（受让）
            '受平手': 0,
            '受平/半': -0.25,
            '受平半': -0.25,
            '受半球': -0.5,
            '受半/一': -0.75,
            '受半一': -0.75,
            '受一球': -1,
            '受一/球半': -1.25,
            '受一球/球半': -1.25,  # v24补充: 别名
            '受一球半': -1.5,
            '受球半': -1.5,
            '受两球': -2,
            '受二球': -2,
        }
        
        # 1. 先尝试中文盘口映射
        if handicap_str in chinese_handicap_map:
            return float(chinese_handicap_map[handicap_str])
        
        # 2. 尝试直接转换为数字
        try:
            return float(handicap_str)
        except ValueError:
            pass
        
        # 3. 处理带/的格式（如"2/2.5" -> 2.25）
        if '/' in handicap_str:
            parts = handicap_str.split('/')
            if len(parts) == 2:
                try:
                    lower = float(parts[0])
                    upper = float(parts[1])
                    return (lower + upper) / 2
                except ValueError:
                    return None
        
        return None

    @staticmethod
    def _extract_ou_row(row_data):
        if len(row_data) >= 5:
            goal_line_raw = row_data[3] if len(row_data) > 3 else ''
            handicap_value = OddsFetcher._convert_handicap_to_number(goal_line_raw)
            return {
                'time': row_data[0] if len(row_data) > 0 else '',
                'score': row_data[1] if len(row_data) > 1 else '',
                'over_odds': row_data[2] if len(row_data) > 2 else '',
                'goal_line': goal_line_raw,
                'handicap': str(handicap_value) if handicap_value is not None else goal_line_raw,
                'under_odds': row_data[4] if len(row_data) > 4 else '',
                'change_time': row_data[5] if len(row_data) > 5 else '',
                'status': row_data[6] if len(row_data) > 6 else '',
            }
        return {}

    @staticmethod
    def _parse_asian_table(html, company_name=''):
        try:
            soup = BeautifulSoup(html, 'html.parser')
            table = None
            odds_span = soup.find('span', id='odds2')
            if odds_span:
                table = odds_span.find('table')
            if not table:
                tables = soup.find_all('table')
                for t in tables:
                    if len(t.find_all('tr')) > 3:
                        table = t
                        break
            if not table:
                return {'header': [], 'rows': [], 'error': '未找到亚盘表格'}
            rows = table.find_all('tr')
            if len(rows) < 2:
                return {'header': [], 'rows': [], 'error': '表格行数不足'}
            header_row = rows[0]
            headers = []
            for td in header_row.find_all(['td', 'th']):
                text = td.get_text(strip=True)
                headers.append(text if text else '')
            if all(h == '' for h in headers):
                headers = ['时间', '比分', '主队', '盘口', '客队', '变化时间', '状态']
            data_rows = []
            for row in rows[1:]:
                tds = row.find_all('td')
                if len(tds) < 5:
                    continue
                row_data = []
                has_data = False
                status_value = ''
                for idx, td in enumerate(tds):
                    b_tag = td.find('b')
                    if b_tag:
                        text = b_tag.get_text(strip=True)
                    else:
                        text = td.get_text(strip=True)
                    td_class = td.get('class', [])
                    td_class_str = ' '.join(td_class) if isinstance(td_class, list) else str(td_class)
                    if 'hg_red' in td_class_str:
                        status_value = '即'
                    elif 'hg_blue' in td_class_str:
                        status_value = '滚'
                    elif 'hg_green' in td_class_str:
                        if text == '封':
                            status_value = '封'
                        elif not status_value:
                            status_value = '未知'
                    row_data.append(text)
                    if text and text != '封':
                        has_data = True
                if row_data and len(row_data) >= 7:
                    if not row_data[-1] or row_data[-1] in ['', ' ']:
                        row_data[-1] = status_value if status_value else ''
                elif row_data and status_value and len(row_data) >= 6:
                    while len(row_data) < 7:
                        row_data.append('')
                    if len(row_data) >= 7:
                        row_data[6] = status_value
                if has_data and row_data:
                    data_rows.append(row_data)
            return {'header': headers, 'rows': data_rows}
        except Exception as e:
            return {'header': [], 'rows': [], 'error': str(e)}

    @staticmethod
    def _parse_overunder_table(html, company_name=''):
        try:
            soup = BeautifulSoup(html, 'html.parser')
            table = None
            odds_span = soup.find('span', id='odds2')
            if odds_span:
                table = odds_span.find('table')
            if not table:
                tables = soup.find_all('table') if hasattr(soup, 'find_all') else []
                for t in tables:
                    if len(t.find_all('tr')) > 3:
                        table = t
                        break
            if not table:
                return {'header': [], 'rows': [], 'error': '未找到大小球表格'}
            rows = table.find_all('tr')
            if len(rows) < 2:
                return {'header': [], 'rows': [], 'error': '表格行数不足'}
            header_row = rows[0]
            headers = []
            for td in header_row.find_all(['td', 'th']):
                text = td.get_text(strip=True)
                headers.append(text if text else '')
            if all(h == '' for h in headers):
                headers = ['时间', '比分', '大球', '进球数', '小球', '变化时间', '状态']
            data_rows = []
            for row in rows[1:]:
                tds = row.find_all('td')
                if len(tds) < 5:
                    continue
                row_data = []
                has_data = False
                status_value = ''
                for idx, td in enumerate(tds):
                    b_tag = td.find('b')
                    if b_tag:
                        text = b_tag.get_text(strip=True)
                    else:
                        text = td.get_text(strip=True)
                    td_class = td.get('class', [])
                    td_class_str = ' '.join(td_class) if isinstance(td_class, list) else str(td_class)
                    if 'hg_red' in td_class_str:
                        status_value = '即'
                    elif 'hg_blue' in td_class_str:
                        status_value = '滚'
                    elif 'hg_green' in td_class_str:
                        if text == '封':
                            status_value = '封'
                        elif not status_value:
                            status_value = '未知'
                    row_data.append(text)
                    if text and text != '封':
                        has_data = True
                if row_data and len(row_data) >= 7:
                    if not row_data[-1] or row_data[-1] in ['', ' ']:
                        row_data[-1] = status_value if status_value else ''
                elif row_data and status_value and len(row_data) >= 6:
                    while len(row_data) < 7:
                        row_data.append('')
                    if len(row_data) >= 7:
                        row_data[6] = status_value
                if has_data and row_data:
                    data_rows.append(row_data)
            return {'header': headers, 'rows': data_rows}
        except Exception as e:
            return {'header': [], 'rows': [], 'error': str(e)}


# ============================================================
# 初盘筛选配置控件
# ============================================================
class InitialOddsFilterWidget(QWidget):
    """单个初盘筛选配置控件（半场/全场亚盘或大小球）"""
    
    # 亚盘盘口选项
    ASIAN_HANDICAP_ITEMS = [
        ('平手', '0'), ('平/半', '0.25'), ('半球', '0.5'), ('半/一', '0.75'),
        ('一球', '1'), ('一/球半', '1.25'), ('球半', '1.5'), ('球半/两', '1.75'),
        ('两球', '2'), ('两/两球半', '2.25'), ('两球半', '2.5'), ('两球半/三', '2.75'),
        ('三球', '3'), ('三/三球半', '3.25'), ('三球半', '3.5'),
    ]
    
    # 大小球盘口选项
    OVERUNDER_ITEMS = [
        ('1.5/2', '1.75'), ('2', '2'), ('2/2.5', '2.25'), ('2.5', '2.5'),
        ('2.5/3', '2.75'), ('3', '3'), ('3/3.5', '3.25'), ('3.5', '3.5'),
        ('3.5/4', '3.75'), ('4', '4'), ('4/4.5', '4.25'), ('4.5', '4.5'),
        ('4.5/5', '4.75'), ('5/5.5', '5.25'), ('6', '6'),
    ]
    
    def __init__(self, title, filter_type='asian', parent=None):
        """
        :param title: 标题（如"半场亚盘初盘"）
        :param filter_type: 筛选类型 'asian' 或 'overunder'
        """
        super().__init__(parent)
        self.filter_type = filter_type
        self._build_ui(title)
    
    def _build_ui(self, title):
        """构建UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)
        
        # 启用复选框
        self.enabled_cb = QCheckBox(f"启用{title}")
        self.enabled_cb.setChecked(False)
        layout.addWidget(self.enabled_cb)
        
        # 参数区域
        param_group = QGroupBox("筛选条件")
        param_layout = QGridLayout(param_group)
        param_layout.setSpacing(8)
        
        # 主队水位范围
        param_layout.addWidget(QLabel("主水:"), 0, 0)
        self.home_min_spin = QDoubleSpinBox()
        self.home_min_spin.setRange(0.01, 10.0)
        self.home_min_spin.setValue(0.80)
        self.home_min_spin.setSingleStep(0.05)
        self.home_min_spin.setDecimals(2)
        self.home_min_spin.setEnabled(False)
        param_layout.addWidget(self.home_min_spin, 0, 1)
        
        param_layout.addWidget(QLabel("~"), 0, 2)
        
        self.home_max_spin = QDoubleSpinBox()
        self.home_max_spin.setRange(0.01, 10.0)
        self.home_max_spin.setValue(1.00)
        self.home_max_spin.setSingleStep(0.05)
        self.home_max_spin.setDecimals(2)
        self.home_max_spin.setEnabled(False)
        param_layout.addWidget(self.home_max_spin, 0, 3)
        
        # 盘口选择
        param_layout.addWidget(QLabel("盘口:"), 1, 0)
        self.handicap_combo = QComboBox()
        self.handicap_combo.addItem("-- 不限 --", "")
        
        if self.filter_type == 'asian':
            for label, value in self.ASIAN_HANDICAP_ITEMS:
                self.handicap_combo.addItem(label, value)
        else:
            for label, value in self.OVERUNDER_ITEMS:
                self.handicap_combo.addItem(label, value)
        
        self.handicap_combo.setEnabled(False)
        param_layout.addWidget(self.handicap_combo, 1, 1, 1, 3)
        
        # 客队水位范围
        param_layout.addWidget(QLabel("客水:"), 2, 0)
        self.away_min_spin = QDoubleSpinBox()
        self.away_min_spin.setRange(0.01, 10.0)
        self.away_min_spin.setValue(0.80)
        self.away_min_spin.setSingleStep(0.05)
        self.away_min_spin.setDecimals(2)
        self.away_min_spin.setEnabled(False)
        param_layout.addWidget(self.away_min_spin, 2, 1)
        
        param_layout.addWidget(QLabel("~"), 2, 2)
        
        self.away_max_spin = QDoubleSpinBox()
        self.away_max_spin.setRange(0.01, 10.0)
        self.away_max_spin.setValue(1.00)
        self.away_max_spin.setSingleStep(0.05)
        self.away_max_spin.setDecimals(2)
        self.away_max_spin.setEnabled(False)
        param_layout.addWidget(self.away_max_spin, 2, 3)
        
        layout.addWidget(param_group)
        
        # 连接信号
        self.enabled_cb.toggled.connect(self._toggle_inputs)
    
    def _toggle_inputs(self, enabled):
        """切换输入框启用状态"""
        self.home_min_spin.setEnabled(enabled)
        self.home_max_spin.setEnabled(enabled)
        self.handicap_combo.setEnabled(enabled)
        self.away_min_spin.setEnabled(enabled)
        self.away_max_spin.setEnabled(enabled)
    
    def get_config(self):
        """获取配置"""
        return {
            'enabled': self.enabled_cb.isChecked(),
            'home_min': self.home_min_spin.value(),
            'home_max': self.home_max_spin.value(),
            'handicap': self.handicap_combo.currentData(),
            'away_min': self.away_min_spin.value(),
            'away_max': self.away_max_spin.value(),
        }
    
    def set_config(self, config):
        """v24新增: 设置配置"""
        if not config:
            return
        
        # 设置启用状态
        self.enabled_cb.setChecked(config.get('enabled', False))
        
        # 设置水位范围
        self.home_min_spin.setValue(config.get('home_min', 0.80))
        self.home_max_spin.setValue(config.get('home_max', 1.00))
        self.away_min_spin.setValue(config.get('away_min', 0.80))
        self.away_max_spin.setValue(config.get('away_max', 1.00))
        
        # 设置盘口
        handicap_value = config.get('handicap', '')
        index = self.handicap_combo.findData(handicap_value)
        if index >= 0:
            self.handicap_combo.setCurrentIndex(index)


# ============================================================
# 主窗口
# ============================================================
class FilterToolMainWindow(QMainWindow):
    """比赛初盘筛选工具主窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("比赛初盘筛选工具 v1.0")
        self.setMinimumSize(1200, 800)
        
        # 初始化组件
        self.filter_ctrl = None
        self.fetcher = None
        self.filtered_matches = []
        self.proxy_config = None  # v24新增: 代理配置
        
        self._init_ui()
        self._load_proxy_config()  # v24修复: 在UI初始化后加载代理配置
        self._init_components()
    
    def _init_ui(self):
        """初始化UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # === 顶部：筛选配置区 ===
        config_scroll = QScrollArea()
        config_scroll.setWidgetResizable(True)
        config_scroll.setMaximumHeight(400)
        
        config_widget = QWidget()
        config_layout = QVBoxLayout(config_widget)
        config_layout.setSpacing(10)
        
        # 四个筛选器
        self.half_asian_filter = InitialOddsFilterWidget("半场亚盘初盘", 'asian')
        self.full_asian_filter = InitialOddsFilterWidget("全场亚盘初盘", 'asian')
        self.half_ou_filter = InitialOddsFilterWidget("半场大小球初盘", 'overunder')
        self.full_ou_filter = InitialOddsFilterWidget("全场大小球初盘", 'overunder')
        
        config_layout.addWidget(self.half_asian_filter)
        config_layout.addWidget(self.full_asian_filter)
        config_layout.addWidget(self.half_ou_filter)
        config_layout.addWidget(self.full_ou_filter)
        
        config_scroll.setWidget(config_widget)
        main_layout.addWidget(config_scroll)
        
        # === 中部：按钮区 ===
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        self.start_btn = QPushButton("▶ 开始筛选")
        self.start_btn.setObjectName("startBtn")
        self.start_btn.clicked.connect(self.on_start_filter)
        btn_layout.addWidget(self.start_btn)
        
        self.export_btn = QPushButton("💾 导出Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.clicked.connect(self.on_export_excel)
        self.export_btn.setEnabled(False)
        btn_layout.addWidget(self.export_btn)
        
        # v24新增: 方案管理按钮
        self.save_profile_btn = QPushButton("📝 保存方案")
        self.save_profile_btn.clicked.connect(self.on_save_profile)
        btn_layout.addWidget(self.save_profile_btn)
        
        self.load_profile_btn = QPushButton("📂 加载方案")
        self.load_profile_btn.clicked.connect(self.on_load_profile)
        btn_layout.addWidget(self.load_profile_btn)
        
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)
        
        # === 底部：结果展示区 ===
        result_group = QGroupBox("筛选结果")
        result_layout = QVBoxLayout(result_group)
        
        # v24修复: 统计信息和按钮
        stats_btn_layout = QHBoxLayout()
        self.stats_label = QLabel("共 0 场比赛")
        self.stats_label.setStyleSheet("font-size: 12px; color: #7f8c8d;")
        stats_btn_layout.addWidget(self.stats_label)
        stats_btn_layout.addStretch()
        
        # v24新增: 打开详细结果窗口按钮
        self.show_result_btn = QPushButton("📊 查看详细结果 (新窗口)")
        self.show_result_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.show_result_btn.clicked.connect(self.show_result_dialog)
        self.show_result_btn.setEnabled(False)  # 初始禁用
        stats_btn_layout.addWidget(self.show_result_btn)
        
        # v25新增: 跳过完赛比赛复选框
        self.skip_finished_cb = QCheckBox("跳过完赛比赛")
        self.skip_finished_cb.setChecked(True)
        self.skip_finished_cb.setToolTip("勾选后自动跳过状态为「完」的比赛")
        stats_btn_layout.addWidget(self.skip_finished_cb)
        
        result_layout.addLayout(stats_btn_layout)
        
        # v24修复: 简化主窗口的表格（仅显示基本信息）
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(6)
        self.result_table.setHorizontalHeaderLabels([
            '联赛', '时间', '主队', '比分', '客队', '比赛ID'
        ])
        
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        
        self.result_table.setMaximumHeight(200)  # 限制主窗口表格高度
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        result_layout.addWidget(self.result_table)
        
        main_layout.addWidget(result_group)
        
        # === 日志区 ===
        log_group = QGroupBox("运行日志")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(150)
        log_layout.addWidget(self.log_text)
        
        main_layout.addWidget(log_group)
        
        # 状态栏
        self.statusBar().showMessage("就绪")
    
    def _init_components(self):
        """初始化组件"""
        self.filter_ctrl = FilterController()
        # v24修复: 使用代理配置创建数据获取器
        self.fetcher = OddsFetcher(proxy_config=self.proxy_config)
    
    def _load_proxy_config(self):
        """v24新增: 加载代理配置文件（支持 macOS .app 打包）"""
        import json
        try:
            # 1. 用户配置目录（可写）
            if sys.platform == 'darwin':
                user_dir = os.path.join(os.path.expanduser("~"), "Library", "Application Support", "比赛初盘筛选工具")
            else:
                user_dir = os.path.dirname(os.path.abspath(__file__))
            user_path = os.path.join(user_dir, "proxy_config.json")
            
            if os.path.exists(user_path):
                with open(user_path, 'r', encoding='utf-8') as f:
                    self.proxy_config = json.load(f)
                self.add_log(f"✅ 已加载代理配置(用户目录)")
                return
            
            # 2. .app 资源目录（打包内置，只读回退）
            try:
                bundle_path = os.path.join(sys._MEIPASS, "proxy_config.json")
                if os.path.exists(bundle_path):
                    with open(bundle_path, 'r', encoding='utf-8') as f:
                        self.proxy_config = json.load(f)
                    self.add_log(f"✅ 已加载代理配置(app资源)")
                    return
            except (AttributeError, ImportError):
                pass
            
            # 3. 上级目录（开发模式）
            parent_path = os.path.join(os.path.dirname(__file__), '..', 'proxy_config.json')
            if os.path.exists(parent_path):
                with open(parent_path, 'r', encoding='utf-8') as f:
                    self.proxy_config = json.load(f)
                self.add_log(f"✅ 已加载代理配置(上级目录)")
                return
            
            # 4. 未找到
            self.proxy_config = {'enabled': False}
            self.add_log("⚠️ 未找到代理配置文件，将不使用代理")
        except Exception as e:
            self.proxy_config = {'enabled': False}
            self.add_log(f"❌ 加载代理配置失败: {e}")
    
    def add_log(self, message):
        """添加日志"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.append(f"[{timestamp}] {message}")
        # 自动滚动到底部
        sb = self.log_text.verticalScrollBar()
        sb.setValue(sb.maximum())
    
    def on_start_filter(self):
        """开始筛选"""
        # 获取所有筛选配置
        configs = {
            'half_asian': self.half_asian_filter.get_config(),
            'full_asian': self.full_asian_filter.get_config(),
            'half_ou': self.half_ou_filter.get_config(),
            'full_ou': self.full_ou_filter.get_config(),
        }
        
        # 检查是否至少启用一个筛选
        if not any(c['enabled'] for c in configs.values()):
            QMessageBox.warning(self, "提示", "请至少启用一个筛选条件！")
            return
        
        self.add_log("=" * 60)
        self.add_log("开始执行筛选...")
        self.start_btn.setEnabled(False)
        self.statusBar().showMessage("筛选中...")
        
        # 在后台线程执行筛选
        skip_finished = self.skip_finished_cb.isChecked()
        self.worker = FilterWorkerThread(self.filter_ctrl, self.fetcher, configs, skip_finished=skip_finished)
        self.worker.finished.connect(self.on_filter_finished)
        self.worker.progress.connect(self.add_log)
        self.worker.start()
    
    def on_filter_finished(self, success, matches, error_msg):
        """筛选完成回调"""
        self.start_btn.setEnabled(True)
        
        if not success:
            QMessageBox.critical(self, "错误", f"筛选失败：{error_msg}")
            self.statusBar().showMessage("筛选失败")
            return
        
        self.filtered_matches = matches
        self.display_results(matches)
        
        self.export_btn.setEnabled(len(matches) > 0)
        self.statusBar().showMessage(f"筛选完成，共 {len(matches)} 场比赛")
        self.add_log(f"✅ 筛选完成！共找到 {len(matches)} 场符合条件的比赛")
    
    def display_results(self, matches):
        """v24修复: 显示筛选结果（主窗口仅显示基本信息）"""
        self.result_table.setRowCount(0)
        
        for match in matches:
            row = self.result_table.rowCount()
            self.result_table.insertRow(row)
            
            # v24修复: 只显示基本信息
            self.result_table.setItem(row, 0, QTableWidgetItem(match.get('league', '')))
            self.result_table.setItem(row, 1, QTableWidgetItem(match.get('match_time', '')))
            self.result_table.setItem(row, 2, QTableWidgetItem(match.get('home_team', '')))
            self.result_table.setItem(row, 3, QTableWidgetItem(match.get('score', '')))
            self.result_table.setItem(row, 4, QTableWidgetItem(match.get('away_team', '')))
            self.result_table.setItem(row, 5, QTableWidgetItem(match.get('match_id', '')))
        
        self.stats_label.setText(f"共 {len(matches)} 场比赛")
        
        # v24新增: 启用查看详细结果按钮
        if len(matches) > 0:
            self.show_result_btn.setEnabled(True)
        else:
            self.show_result_btn.setEnabled(False)
    
    def on_export_excel(self):
        """导出Excel"""
        if not self.filtered_matches:
            QMessageBox.warning(self, "提示", "没有可导出的数据！")
            return
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 构建筛选条件描述
        conditions = []
        if self.half_asian_filter.enabled_cb.isChecked():
            conditions.append("半场亚盘")
        if self.full_asian_filter.enabled_cb.isChecked():
            conditions.append("全场亚盘")
        if self.half_ou_filter.enabled_cb.isChecked():
            conditions.append("半场大小球")
        if self.full_ou_filter.enabled_cb.isChecked():
            conditions.append("全场大小球")
        
        condition_str = "_".join(conditions) if conditions else "筛选"
        default_filename = f"比赛筛选_{condition_str}_{timestamp}.xlsx"
        
        # 选择保存路径
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "保存Excel文件",
            default_filename,
            "Excel文件 (*.xlsx);;所有文件 (*)"
        )
        
        if not filepath:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "筛选结果"
            
            # 设置表头（v23优化：增加半场和全场初盘数据）
            headers = [
                '联赛', '时间', '状态', '主队', '比分', '客队',
                '全场亚盘初盘', '全场大小球初盘',
                '半场亚盘初盘', '半场大小球初盘',
                '比赛ID'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # v23新增: 辅助函数 - 格式化初盘数据
            def format_asian_initial(data):
                """格式化亚盘初盘"""
                if not data:
                    return "-"
                home = data.get('home_odds', '-')
                handicap = data.get('handicap', '-')
                away = data.get('away_odds', '-')
                return f"{home} {handicap} {away}"
            
            def format_ou_initial(data):
                """格式化大小球初盘"""
                if not data:
                    return "-"
                over = data.get('over_odds', '-')
                goal_line = data.get('goal_line', '-')
                under = data.get('under_odds', '-')
                return f"{over} {goal_line} {under}"
            
            # 填充数据
            for row_idx, match in enumerate(self.filtered_matches, 2):
                ws.cell(row=row_idx, column=1, value=match.get('league', ''))
                ws.cell(row=row_idx, column=2, value=match.get('match_time', ''))
                ws.cell(row=row_idx, column=3, value=match.get('status', ''))
                ws.cell(row=row_idx, column=4, value=match.get('home_team', ''))
                ws.cell(row=row_idx, column=5, value=match.get('score', ''))
                ws.cell(row=row_idx, column=6, value=match.get('away_team', ''))
                
                # v23新增: 全场初盘
                ws.cell(row=row_idx, column=7, value=format_asian_initial(match.get('full_asian_initial')))
                ws.cell(row=row_idx, column=8, value=format_ou_initial(match.get('full_ou_initial')))
                
                # v23新增: 半场初盘
                ws.cell(row=row_idx, column=9, value=format_asian_initial(match.get('half_asian_initial')))
                ws.cell(row=row_idx, column=10, value=format_ou_initial(match.get('half_ou_initial')))
                
                ws.cell(row=row_idx, column=11, value=match.get('match_id', ''))
            
            # 调整列宽
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 25  # 全场亚盘
            ws.column_dimensions['H'].width = 25  # 全场大小球
            ws.column_dimensions['I'].width = 25  # 半场亚盘
            ws.column_dimensions['J'].width = 25  # 半场大小球
            ws.column_dimensions['K'].width = 15  # 比赛ID
            
            wb.save(filepath)
            
            QMessageBox.information(self, "成功", f"文件已保存到：\n{filepath}")
            self.add_log(f"✅ Excel文件已导出：{filepath}")
            
        except ImportError:
            QMessageBox.critical(self, "错误", "未安装openpyxl库！\n请运行：pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{e}")
            self.add_log(f"❌ 导出失败：{e}")
    
    def show_result_dialog(self):
        """v24新增: 显示详细结果对话框（新窗口）"""
        if not self.filtered_matches:
            QMessageBox.warning(self, "提示", "没有可显示的数据！")
            return
        
        # 创建对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("📊 筛选结果详情")
        dialog.setMinimumSize(1400, 800)  # v24修复: 设置较大的默认尺寸
        
        layout = QVBoxLayout(dialog)
        
        # 统计信息
        stats_label = QLabel(f"共 {len(self.filtered_matches)} 场比赛 | 双击行可复制比赛ID")
        stats_label.setStyleSheet("font-size: 13px; color: #2c3e50; padding: 5px;")
        layout.addWidget(stats_label)
        
        # 表格
        table = QTableWidget()
        table.setColumnCount(11)
        table.setHorizontalHeaderLabels([
            '联赛', '时间', '状态', '主队', '比分', '客队',
            '全场亚盘初盘', '全场大小球初盘',
            '半场亚盘初盘', '半场大小球初盘',
            '比赛ID'
        ])
        
        # v24修复: 设置列宽策略
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 联赛
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 时间
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 状态
        header.setSectionResizeMode(3, QHeaderView.Stretch)           # 主队
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # 比分
        header.setSectionResizeMode(5, QHeaderView.Stretch)           # 客队
        header.setSectionResizeMode(6, QHeaderView.Interactive)       # 全场亚盘
        header.setSectionResizeMode(7, QHeaderView.Interactive)       # 全场大小球
        header.setSectionResizeMode(8, QHeaderView.Interactive)       # 半场亚盘
        header.setSectionResizeMode(9, QHeaderView.Interactive)       # 半场大小球
        header.setSectionResizeMode(10, QHeaderView.ResizeToContents) # 比赛ID
        
        # 设置初始列宽
        table.setColumnWidth(6, 180)  # 全场亚盘
        table.setColumnWidth(7, 180)  # 全场大小球
        table.setColumnWidth(8, 180)  # 半场亚盘
        table.setColumnWidth(9, 180)  # 半场大小球
        
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # v24新增: 格式化函数
        def format_asian_initial(data):
            """格式化亚盘初盘"""
            if not data:
                return "-"
            home = data.get('home_odds', '-')
            handicap = data.get('handicap', '-')
            away = data.get('away_odds', '-')
            return f"{home} {handicap} {away}"
        
        def format_ou_initial(data):
            """格式化大小球初盘"""
            if not data:
                return "-"
            over = data.get('over_odds', '-')
            goal_line = data.get('goal_line', '-')
            under = data.get('under_odds', '-')
            return f"{over} {goal_line} {under}"
        
        # 填充数据
        for match in self.filtered_matches:
            row = table.rowCount()
            table.insertRow(row)
            
            table.setItem(row, 0, QTableWidgetItem(match.get('league', '')))
            table.setItem(row, 1, QTableWidgetItem(match.get('match_time', '')))
            table.setItem(row, 2, QTableWidgetItem(match.get('status', '')))
            table.setItem(row, 3, QTableWidgetItem(match.get('home_team', '')))
            table.setItem(row, 4, QTableWidgetItem(match.get('score', '')))
            table.setItem(row, 5, QTableWidgetItem(match.get('away_team', '')))
            
            # 全场初盘
            table.setItem(row, 6, QTableWidgetItem(format_asian_initial(match.get('full_asian_initial'))))
            table.setItem(row, 7, QTableWidgetItem(format_ou_initial(match.get('full_ou_initial'))))
            
            # 半场初盘
            table.setItem(row, 8, QTableWidgetItem(format_asian_initial(match.get('half_asian_initial'))))
            table.setItem(row, 9, QTableWidgetItem(format_ou_initial(match.get('half_ou_initial'))))
            
            table.setItem(row, 10, QTableWidgetItem(match.get('match_id', '')))
        
        layout.addWidget(table)
        
        # 按钮
        btn_layout = QHBoxLayout()
        
        copy_btn = QPushButton("📋 复制选中比赛ID")
        copy_btn.clicked.connect(lambda: self.copy_selected_match_id(table))
        btn_layout.addWidget(copy_btn)
        
        btn_layout.addStretch()
        
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        # 显示对话框
        dialog.exec_()
    
    def copy_selected_match_id(self, table):
        """v24新增: 复制选中的比赛ID"""
        selected_rows = table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "提示", "请先选择一行！")
            return
        
        # 获取第一行的比赛ID（第11列，索引10）
        row = selected_rows[0].row()
        match_id_item = table.item(row, 10)
        if match_id_item:
            match_id = match_id_item.text()
            clipboard = QApplication.clipboard()
            clipboard.setText(match_id)
            QMessageBox.information(self, "成功", f"已复制比赛ID：{match_id}")
        else:
            QMessageBox.warning(self, "提示", "未找到比赛ID！")
    
    def on_save_profile(self):
        """v24新增: 保存筛选方案（跨平台兼容）"""
        import json
        # 获取当前配置
        profile_data = {
            'half_asian': self.half_asian_filter.get_config(),
            'full_asian': self.full_asian_filter.get_config(),
            'half_ou': self.half_ou_filter.get_config(),
            'full_ou': self.full_ou_filter.get_config(),
        }
        
        # 选择保存路径
        # mac: 默认保存到桌面，避免根目录只读问题
        if sys.platform == 'darwin':
            default_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        else:
            default_dir = os.path.expanduser("~")
        default_name = f"筛选方案_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        default_path = os.path.join(default_dir, default_name)
        
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "保存筛选方案",
            default_path,
            "JSON文件 (*.json);;所有文件 (*)",
        )
        
        if not filepath:
            return
        
        # 确保文件名有 .json 后缀且不含逗号等非法字符
        base, ext = os.path.splitext(filepath)
        base = base.replace(',', '_').replace('，', '_')
        if not ext or ext.lower() != '.json':
            ext = '.json'
        filepath = base + ext
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(profile_data, f, ensure_ascii=False, indent=2)
            
            QMessageBox.information(self, "成功", f"方案已保存到：\n{filepath}")
            self.add_log(f"✅ 方案已保存：{filepath}")
        except Exception as e:
            self.add_log(f"❌ 保存方案失败：{type(e).__name__}: {e}")
            QMessageBox.critical(self, "错误", f"保存失败：{e}")
    def on_load_profile(self):
        """v24新增: 加载筛选方案（跨平台兼容）"""
        import json
        # 选择加载文件（macOS使用非原生对话框避免兼容性问题）
        options = QFileDialog.Options()
        if sys.platform == 'darwin':
            options |= QFileDialog.DontUseNativeDialog
        
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "加载筛选方案",
            os.path.expanduser("~"),
            "JSON文件 (*.json);;所有文件 (*)",
            options=options
        )
        
        if not filepath:
            return
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                profile_data = json.load(f)
            
            # 应用配置到各个筛选器
            if 'half_asian' in profile_data:
                self.half_asian_filter.set_config(profile_data['half_asian'])
            if 'full_asian' in profile_data:
                self.full_asian_filter.set_config(profile_data['full_asian'])
            if 'half_ou' in profile_data:
                self.half_ou_filter.set_config(profile_data['half_ou'])
            if 'full_ou' in profile_data:
                self.full_ou_filter.set_config(profile_data['full_ou'])
            
            QMessageBox.information(self, "成功", f"方案已加载：\n{filepath}")
            self.add_log(f"✅ 方案已加载：{filepath}")
        except Exception as e:
            self.add_log(f"❌ 加载方案失败：{type(e).__name__}: {e}")
            QMessageBox.critical(self, "错误", f"加载失败：{e}")


# ============================================================
# 筛选工作线程
# ============================================================
class FilterWorkerThread(QThread):
    """后台筛选工作线程"""
    
    finished = pyqtSignal(bool, list, str)  # (success, matches, error_msg)
    progress = pyqtSignal(str)
    
    def __init__(self, filter_ctrl, fetcher, configs, skip_finished=False):
        super().__init__()
        self.filter_ctrl = filter_ctrl
        self.fetcher = fetcher
        self.configs = configs
        self.skip_finished = skip_finished
    
    def run(self):
        """执行筛选"""
        try:
            # 步骤1：使用浏览器筛选获取比赛列表
            self.progress.emit("步骤1: 启动浏览器并进行盘口筛选...")
            
            # v24修复: 确定需要哪些浏览器筛选（只考虑全场，半场通过API本地筛选）
            asian_enabled = self.configs['full_asian']['enabled']  # 只看全场亚盘
            ou_enabled = self.configs['full_ou']['enabled']        # 只看全场大小球
            
            # v24修复: 收集全场盘口值（用于浏览器筛选）
            asian_values = []
            asian_need_browser_filter = False
            asian_need_get_all = False
            
            if self.configs['full_asian']['enabled']:
                val = self.configs['full_asian']['handicap']
                if val:  # 如果选择了具体盘口
                    asian_values.append(val)
                    asian_need_browser_filter = True
                else:  # 如果选择"不限"
                    asian_need_get_all = True
            
            ou_values = []
            ou_need_browser_filter = False
            ou_need_get_all = False
            
            if self.configs['full_ou']['enabled']:
                val = self.configs['full_ou']['handicap']
                if val:
                    if val not in ou_values:
                        ou_values.append(val)
                    ou_need_browser_filter = True
                else:
                    ou_need_get_all = True
            
            # v24修复: 判断是否需要获取所有比赛
            # 如果启用了半场筛选但没有启用全场筛选，需要获取所有比赛
            need_get_all_matches = (
                (self.configs['half_asian']['enabled'] or self.configs['half_ou']['enabled']) and  # 有半场筛选
                not asian_need_browser_filter and not ou_need_browser_filter  # 但没有全场浏览器筛选
            )
            
            need_browser_filter = asian_need_browser_filter or ou_need_browser_filter
            
            # v22修复: 设置日志回调，通过线程的progress信号安全输出日志
            self.filter_ctrl.set_log_callback(lambda msg: self.progress.emit(msg))
            self.filter_ctrl.open_live_page()
            
            if need_get_all_matches and not need_browser_filter:
                # 如果只启用了"不限"，没有指定具体盘口，直接获取所有比赛
                self.progress.emit("步骤1: 获取所有比赛（不限盘口）...")
                matches = self.filter_ctrl.get_all_matches(exclude_finished=self.skip_finished)
            elif need_browser_filter:
                # 如果有指定具体盘口，执行浏览器筛选
                self.progress.emit("步骤1: 启动浏览器并进行盘口筛选...")
                matches = self.filter_ctrl.filter_combined(
                    asian_enabled=asian_need_browser_filter,
                    asian_values=asian_values,
                    ou_enabled=ou_need_browser_filter,
                    ou_values=ou_values,
                    skip_finished=self.skip_finished
                )
            else:
                # 理论上不会到这里，因为前面已经检查过至少启用一个筛选
                matches = self.filter_ctrl.get_all_matches(exclude_finished=self.skip_finished)
            
            if not matches:
                self.finished.emit(False, [], "未找到任何比赛")
                return
            
            self.progress.emit(f"步骤1完成: 获取到 {len(matches)} 场比赛")
            
            # 步骤2：获取每场比赛的初盘数据并筛选（v23优化：多进程并发+重试机制）
            self.progress.emit("步骤2: 获取初盘数据并筛选...")
            
            filtered_matches = []
            total = len(matches)
            
            # v23新增: 使用多线程并发获取初盘数据（20个并发）
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            def fetch_match_data_with_retry(match_info, max_retries=3):
                """v23新增: 带重试机制的比赛数据获取"""
                match_id = match_info.get('match_id', '')
                last_error = None
                
                for attempt in range(max_retries):
                    try:
                        # 获取全场初盘
                        full_initial = self.fetcher.fetch_initial_odds(match_id)
                        
                        # 获取半场初盘
                        half_initial = self.fetcher.fetch_half_time_initial(match_id)
                        
                        # 合并数据
                        combined_data = {
                            'asian_initial': full_initial.get('asian_initial'),
                            'ou_initial': full_initial.get('ou_initial'),
                            'asian_half_initial': half_initial.get('asian_half_initial'),
                            'ou_half_initial': half_initial.get('ou_half_initial'),
                            'error': full_initial.get('error') or half_initial.get('error'),
                        }
                        
                        if not combined_data['error']:
                            return match_id, combined_data, None
                        
                        last_error = combined_data['error']
                        if attempt < max_retries - 1:
                            time.sleep(random.uniform(0.5, 1.0))  # 重试前等待
                            
                    except Exception as e:
                        last_error = str(e)
                        if attempt < max_retries - 1:
                            time.sleep(random.uniform(0.5, 1.0))
                
                return match_id, None, f'重试{max_retries}次后仍失败: {last_error}'
            
            # v23新增: 使用线程池并发获取（20个并发）
            max_workers = min(20, len(matches))
            success_count = 0
            fail_count = 0
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # 提交所有任务
                future_to_match = {
                    executor.submit(fetch_match_data_with_retry, match): match
                    for match in matches
                }
                
                # 处理完成的任务
                for idx, future in enumerate(as_completed(future_to_match), 1):
                    if not self.isRunning():
                        break
                    
                    match = future_to_match[future]
                    match_id = match.get('match_id', '')
                    
                    try:
                        mid, initial_data, error = future.result(timeout=60)
                        
                        if error:
                            fail_count += 1
                            self.progress.emit(f"  ❌ [{idx}/{total}] {match.get('home_team', '')} vs {match.get('away_team', '')}: {error}")
                            continue
                        
                        success_count += 1
                        
                        # v23新增: 将初盘数据保存到match对象中，供Excel导出使用
                        match['full_asian_initial'] = initial_data.get('asian_initial')
                        match['full_ou_initial'] = initial_data.get('ou_initial')
                        match['half_asian_initial'] = initial_data.get('asian_half_initial')
                        match['half_ou_initial'] = initial_data.get('ou_half_initial')
                        
                        # 检查是否符合所有启用的筛选条件
                        result, reason = self._check_filters(match, initial_data)
                        if result:
                            filtered_matches.append(match)
                            self.progress.emit(f"  ✅ [{idx}/{total}] {match.get('home_team', '')} vs {match.get('away_team', '')}: 符合条件")
                        else:
                            self.progress.emit(f"  ⚪ [{idx}/{total}] {match.get('home_team', '')} vs {match.get('away_team', '')}: 不符合 - {reason}")
                            
                    except Exception as e:
                        fail_count += 1
                        self.progress.emit(f"  ❌ [{idx}/{total}] {match.get('home_team', '')} vs {match.get('away_team', '')}: 异常: {str(e)}")
            
            self.progress.emit(f"步骤2完成: 成功{success_count}场, 失败{fail_count}场, 筛选出 {len(filtered_matches)} 场比赛")
            
            # v24修复: 筛选完成后关闭浏览器
            self.progress.emit("正在关闭浏览器...")
            try:
                self.filter_ctrl.close_browser()
                self.progress.emit("✅ 浏览器已关闭")
            except Exception as e:
                self.progress.emit(f"⚠️ 关闭浏览器失败: {e}")
            
            self.finished.emit(True, filtered_matches, "")
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.progress.emit(f"❌ 筛选异常: {e}")
            
            # v24修复: 异常时也要关闭浏览器
            try:
                self.filter_ctrl.close_browser()
                self.progress.emit("✅ 浏览器已关闭")
            except Exception as close_err:
                self.progress.emit(f"⚠️ 关闭浏览器失败: {close_err}")
            
            self.finished.emit(False, [], str(e))
    
    def _check_filters(self, match, initial_data):
        """v24修复: 检查是否符合所有筛选条件，返回(结果, 原因)"""
        # v23新增: 获取半场和全场数据
        asian_initial = initial_data.get('asian_initial')  # 全场亚盘
        ou_initial = initial_data.get('ou_initial')  # 全场大小球
        asian_half_initial = initial_data.get('asian_half_initial')  # 半场亚盘
        ou_half_initial = initial_data.get('ou_half_initial')  # 半场大小球
        
        # 检查半场亚盘
        if self.configs['half_asian']['enabled']:
            result, reason = self._check_asian_filter(asian_half_initial, self.configs['half_asian'], '半场亚盘')
            if not result:
                return False, reason
        
        # 检查全场亚盘
        if self.configs['full_asian']['enabled']:
            result, reason = self._check_asian_filter(asian_initial, self.configs['full_asian'], '全场亚盘')
            if not result:
                return False, reason
        
        # 检查半场大小球
        if self.configs['half_ou']['enabled']:
            result, reason = self._check_ou_filter(ou_half_initial, self.configs['half_ou'], '半场大小球')
            if not result:
                return False, reason
        
        # 检查全场大小球
        if self.configs['full_ou']['enabled']:
            result, reason = self._check_ou_filter(ou_initial, self.configs['full_ou'], '全场大小球')
            if not result:
                return False, reason
        
        return True, ""
    
    def _check_asian_filter(self, asian_data, config, filter_name='亚盘'):
        """v24修复: 检查亚盘筛选条件，返回(结果, 原因)"""
        if not asian_data:
            return False, f"{filter_name}: 无数据"
        
        home_odds = self._safe_float(asian_data.get('home_odds', ''))
        away_odds = self._safe_float(asian_data.get('away_odds', ''))
        handicap = asian_data.get('handicap', '')
        
        # 检查主水范围
        if home_odds is None:
            return False, f"{filter_name}: 主水无效"
        if not (config['home_min'] <= home_odds <= config['home_max']):
            return False, f"{filter_name}: 主水{home_odds}不在{config['home_min']}-{config['home_max']}范围内"
        
        # 检查客水范围
        if away_odds is None:
            return False, f"{filter_name}: 客水无效"
        if not (config['away_min'] <= away_odds <= config['away_max']):
            return False, f"{filter_name}: 客水{away_odds}不在{config['away_min']}-{config['away_max']}范围内"
        
        # v24修复: 检查盘口（需要转换后比较）
        if config['handicap']:
            # 将网页获取的盘口转换为数字
            handicap_numeric = OddsFetcher._convert_handicap_to_number(handicap)
            # 将配置的盘口值转换为数字
            config_handicap_numeric = OddsFetcher._convert_handicap_to_number(config['handicap'])
            
            # 如果转换失败，使用原始字符串比较
            if handicap_numeric is None or config_handicap_numeric is None:
                if handicap != config['handicap']:
                    return False, f"{filter_name}: 盘口{handicap}不符合{config['handicap']}"
            else:
                # 使用数字比较（允许小误差）
                if abs(handicap_numeric - config_handicap_numeric) > 0.01:
                    return False, f"{filter_name}: 盘口{handicap}({handicap_numeric})不符合{config['handicap']}({config_handicap_numeric})"
        
        return True, ""
    
    def _check_ou_filter(self, ou_data, config, filter_name='大小球'):
        """v24修复: 检查大小球筛选条件，返回(结果, 原因)"""
        if not ou_data:
            return False, f"{filter_name}: 无数据"
        
        over_odds = self._safe_float(ou_data.get('over_odds', ''))
        under_odds = self._safe_float(ou_data.get('under_odds', ''))
        goal_line = ou_data.get('goal_line', '')
        
        # 检查大球水位范围（作为主水）
        if over_odds is None:
            return False, f"{filter_name}: 大球水位无效"
        if not (config['home_min'] <= over_odds <= config['home_max']):
            return False, f"{filter_name}: 大球水位{over_odds}不在{config['home_min']}-{config['home_max']}范围内"
        
        # 检查小球水位范围（作为客水）
        if under_odds is None:
            return False, f"{filter_name}: 小球水位无效"
        if not (config['away_min'] <= under_odds <= config['away_max']):
            return False, f"{filter_name}: 小球水位{under_odds}不在{config['away_min']}-{config['away_max']}范围内"
        
        # v24修复: 检查盘口（需要转换后比较）
        if config['handicap']:
            # 将网页获取的盘口转换为数字
            goal_line_numeric = OddsFetcher._convert_handicap_to_number(goal_line)
            # 将配置的盘口值转换为数字
            config_handicap_numeric = OddsFetcher._convert_handicap_to_number(config['handicap'])
            
            # 如果转换失败，使用原始字符串比较
            if goal_line_numeric is None or config_handicap_numeric is None:
                if goal_line != config['handicap']:
                    return False, f"{filter_name}: 盘口{goal_line}不符合{config['handicap']}"
            else:
                # 使用数字比较（允许小误差）
                if abs(goal_line_numeric - config_handicap_numeric) > 0.01:
                    return False, f"{filter_name}: 盘口{goal_line}({goal_line_numeric})不符合{config['handicap']}({config_handicap_numeric})"
        
        return True, ""
    
    @staticmethod
    def _safe_float(value_str):
        """安全转换为float"""
        try:
            if not value_str or value_str in ['', '-', '封']:
                return None
            return float(value_str)
        except (ValueError, TypeError):
            return None


# ============================================================
# 主程序入口
# ============================================================
def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(GLOBAL_STYLESHEET)
    
    window = FilterToolMainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
